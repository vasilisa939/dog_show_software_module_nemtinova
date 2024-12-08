import openpyxl
import os
import sys
import sqlite3
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QLineEdit, QMessageBox, QComboBox, QListWidget, QListWidgetItem, QTextEdit
)
from PyQt6.QtWidgets import QDialog, QDateEdit
from PyQt6.QtGui import QPixmap

# Создание или подключение к базе данных
conn = sqlite3.connect('dog_show_nemtinova.db')
cursor = conn.cursor()


conn.commit()
def setup_database():
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS Organizer (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        phone TEXT UNIQUE,
        email TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL
    )""")

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS Dog (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        age INTEGER,
        gender TEXT,
        breed_id INTEGER,
        owner_id INTEGER,
        FOREIGN KEY (breed_id) REFERENCES Breed (id),
        FOREIGN KEY (owner_id) REFERENCES Owner (id)
    )""")

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS Breed (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        country TEXT NOT NULL,
        size TEXT NOT NULL
    )""")

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS Owner (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL
    )""")

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS Judge (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        post TEXT NOT NULL
    )""")

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS Show (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        date TEXT NOT NULL,
        location TEXT NOT NULL,
        organizer_id INTEGER,
        judge_id INTEGER,
        dog_id INTEGER,
        FOREIGN KEY (organizer_id) REFERENCES Organizer (id),
        FOREIGN KEY (judge_id) REFERENCES Judge (id),
        FOREIGN KEY (dog_id) REFERENCES Dog (id)
    )""")

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS Result (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        show_id INTEGER,
        dog_id INTEGER,
        score INTEGER,
        FOREIGN KEY (show_id) REFERENCES Show (id),
        FOREIGN KEY (dog_id) REFERENCES Dog (id)
    )""")

    conn.commit()

    cursor.execute("SELECT COUNT(*) FROM Breed")
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO Breed (id, name, country, size) VALUES (1, 'Лабрадор', 'США', 'Большой')")
        cursor.execute("INSERT INTO Breed (id, name, country, size) VALUES (2, 'Пудель', 'Франция', 'Средний')")
        cursor.execute("INSERT INTO Breed (id, name, country, size) VALUES (3, 'Бигль', 'Великобритания', 'Средний')")
        cursor.execute("INSERT INTO Breed (id, name, country, size) VALUES (4, 'Далматин', 'Хорватия', 'Большой')")
        cursor.execute("INSERT INTO Breed (id, name, country, size) VALUES (5, 'Пекинес', 'Китай', 'Маленький')")
        cursor.execute("INSERT INTO Breed (id, name, country, size) VALUES (6, 'Овчарка', 'Германия', 'Большой')")

    cursor.execute("SELECT COUNT(*) FROM Owner")
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO Owner (id, name) VALUES (1, 'Алексей Захаров')")
        cursor.execute("INSERT INTO Owner (id, name) VALUES (2, 'Мария Захарова')")
        cursor.execute("INSERT INTO Owner (id, name) VALUES (3, 'Джейн Доу')")
        cursor.execute("INSERT INTO Owner (id, name) VALUES (4, 'Роман Иванов')")
        cursor.execute("INSERT INTO Owner (id, name) VALUES (5, 'Екатерина Петрова')")


    cursor.execute("SELECT COUNT(*) FROM Judge")
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO Judge (id, name, post) VALUES (1, 'Коротков П.Г.', 'Главный судья')")
        cursor.execute("INSERT INTO Judge (id, name, post) VALUES (2, 'Комарова Р.Л.', 'Судья')")

    cursor.execute("SELECT COUNT(*) FROM Dog")
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO Dog (id, name, age, gender, breed_id, owner_id) VALUES (1, 'Бобик', 3, 'Мальчик', 1, 1)")
        cursor.execute("INSERT INTO Dog (id, name, age, gender, breed_id, owner_id) VALUES (2, 'Шарик', 5, 'Мальчик', 2, 2)")
        cursor.execute("INSERT INTO Dog (id, name, age, gender, breed_id, owner_id) VALUES (3, 'Кексик', 8, 'Мальчик', 3, 3)")
        cursor.execute("INSERT INTO Dog (id, name, age, gender, breed_id, owner_id) VALUES (4, 'Орео', 2, 'Девочка', 4, 4)")
        cursor.execute("INSERT INTO Dog (id, name, age, gender, breed_id, owner_id) VALUES (5, 'Терминатор', 1, 'Девочка', 5, 5)")
        cursor.execute("INSERT INTO Dog (id, name, age, gender, breed_id, owner_id) VALUES (6, 'Тюленчик', 0.5, 'Девочка', 5, 5)")
        cursor.execute("INSERT INTO Dog (id, name, age, gender, breed_id, owner_id) VALUES (7, 'Пупсик', 1, 'Мальчик', 6, 1)")
        cursor.execute("INSERT INTO Dog (id, name, age, gender, breed_id, owner_id) VALUES (8, 'Киви', 9, 'Девочка', 1, 4)")
        cursor.execute("INSERT INTO Dog (id, name, age, gender, breed_id, owner_id) VALUES (9, 'Булка', 5, 'Девочка', 3, 1)")


    conn.commit()


# Основной класс приложения
class AuthWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Авторизация")
        self.setGeometry(300, 300, 400, 200)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("Email")
        layout.addWidget(self.email_input)

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Пароль")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.password_input)

        btn_login = QPushButton("Войти")
        btn_login.clicked.connect(self.login)
        layout.addWidget(btn_login)

        btn_register = QPushButton("Регистрация")
        btn_register.clicked.connect(self.register)
        layout.addWidget(btn_register)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def login(self):
        email = self.email_input.text()
        password = self.password_input.text()

        cursor.execute("SELECT name FROM Organizer WHERE email = ? AND password = ?", (email, password))
        user = cursor.fetchone()

        if user:
            QMessageBox.information(self, "Успех", f"Добро пожаловать, {user[0]}!")
            self.main_menu = MainMenu(user[0])
            self.main_menu.show()
            self.close()
        else:
            QMessageBox.warning(self, "Ошибка", "Неверные данные для входа!")

    def register(self):
        self.registration_window = RegistrationWindow()
        self.registration_window.show()


class RegistrationWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Регистрация")
        self.setGeometry(300, 300, 400, 300)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Имя")
        layout.addWidget(self.name_input)

        self.phone_input = QLineEdit()
        self.phone_input.setPlaceholderText("Телефон")
        layout.addWidget(self.phone_input)

        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("Email")
        layout.addWidget(self.email_input)

        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Пароль")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.password_input)

        btn_register = QPushButton("Зарегистрироваться")
        btn_register.clicked.connect(self.register)
        layout.addWidget(btn_register)

        self.setLayout(layout)

    def register(self):
        name = self.name_input.text()
        phone = self.phone_input.text()
        email = self.email_input.text()
        password = self.password_input.text()

        try:
            cursor.execute("INSERT INTO Organizer (name, phone, email, password) VALUES (?, ?, ?, ?)",
                           (name, phone, email, password))
            conn.commit()
            QMessageBox.information(self, "Успех", "Регистрация прошла успешно!")
            self.close()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Ошибка", "Email или телефон уже используются!")


class MainMenu(QMainWindow):
    def __init__(self, username):
        super().__init__()
        self.username = username
        self.setWindowTitle("Главное меню")
        self.setGeometry(300, 300, 500, 400)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # Загрузка изображения
        image_label = QLabel(self)
        pixmap = QPixmap('dog_mainmenu.jpg')  # Путь к изображению в корне проекта
        image_label.setPixmap(pixmap)
        image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(image_label)

        welcome_label = QLabel()
        welcome_label = QLabel()
        welcome_label.setText(
            f"<html><head/><body><p><span style='font-size:22pt; font-weight:600;'>Добро пожаловать, {self.username} :)</span></p></body></html>")
        layout.addWidget(welcome_label)

        description_label = QLabel(self)
        description_label.setText(
            "Выставка собак — это конкурс красоты, на котором специально обученные и аттестованные эксперты оценивают собак определённой породы в соответствии со стандартом и выбирают лучших.")
        description_label.setWordWrap(True)  # Позволяет тексту переноситься на следующую строку при необходимости
        layout.addWidget(description_label)


        btn_list_dogs = QPushButton("Список собак")
        btn_list_dogs.clicked.connect(self.show_dog_list)
        layout.addWidget(btn_list_dogs)

        btn_create_show = QPushButton("Создать выставку")
        btn_create_show.clicked.connect(self.open_create_show_window)  # Подключаем слот
        layout.addWidget(btn_create_show)

        btn_my_shows = QPushButton("Выставки")
        btn_my_shows.clicked.connect(self.show_show_list)  # Добавляем обработчик клика
        layout.addWidget(btn_my_shows)

        btn_enter_results = QPushButton("Выставить результаты")
        btn_enter_results.clicked.connect(self.enter_results)
        layout.addWidget(btn_enter_results)

        btn_results = QPushButton("Результаты выставок")
        btn_results.clicked.connect(self.show_results_list)
        layout.addWidget(btn_results)

        # Добавляем кнопку выхода
        btn_exit = QPushButton("Выход")
        btn_exit.clicked.connect(self.exit_application)
        layout.addWidget(btn_exit)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def show_dog_list(self):
        self.dog_list_window = DogListWindow()
        self.dog_list_window.show()

    def open_create_show_window(self):
        self.create_show_window = CreateShowWindow()
        self.create_show_window.exec()

    def show_show_list(self):
        self.show_list_window = ShowListWindow()
        self.show_list_window.show()

    def show_results_list(self):
        self.results_window = ShowResultsWindow()
        self.results_window.show()

    def enter_results(self):
        self.enter_results_window = EnterResultsWindow()
        self.enter_results_window.exec()

    def exit_application(self):
        QApplication.instance().quit()


class DogListWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Список собак")
        self.setGeometry(300, 300, 600, 600)

        layout = QVBoxLayout()

        self.dog_info = QTextEdit()
        self.dog_info.setReadOnly(True)
        layout.addWidget(self.dog_info)

        # Виджет для отображения изображения собаки
        self.dog_image_label = QLabel()
        self.dog_image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.dog_image_label)

        # Выпадающее меню для фильтрации по породам
        self.breed_filter = QComboBox()
        self.breed_filter.addItem("Все породы")
        cursor.execute("SELECT name FROM Breed")
        for row in cursor.fetchall():
            self.breed_filter.addItem(row[0])
        self.breed_filter.currentIndexChanged.connect(self.load_dogs)
        layout.addWidget(self.breed_filter)

        self.dog_list = QListWidget()
        self.load_dogs()
        layout.addWidget(self.dog_list)
        self.dog_list.itemClicked.connect(self.display_dog_info)

        self.setLayout(layout)

    def load_dogs(self):
        self.dog_list.clear()
        breed = self.breed_filter.currentText()

        if breed == "Все породы":
            cursor.execute("SELECT Dog.name, Breed.name FROM Dog JOIN Breed ON Dog.breed_id = Breed.id")
        else:
            cursor.execute("""
                SELECT Dog.name, Breed.name FROM Dog
                JOIN Breed ON Dog.breed_id = Breed.id
                WHERE Breed.name = ?
            """, (breed,))

        for row in cursor.fetchall():
            self.dog_list.addItem(f"{row[0]} ({row[1]})")

    def display_dog_info(self, item):
        """Отображает информацию о выбранной собаке и загружает фото."""
        dog_name = item.text().split(' (')[0]

        # Извлекаем всю информацию о собаке, включая её id
        cursor.execute("""
            SELECT Dog.id, Dog.name, Dog.age, Dog.gender, Breed.name, Dog.owner_id
            FROM Dog
            JOIN Breed ON Dog.breed_id = Breed.id
            WHERE Dog.name = ?
        """, (dog_name,))

        dog_info = cursor.fetchone()

        if dog_info:
            info_text = (
                f"Имя: {dog_info[1]}\n"
                f"Возраст: {dog_info[2]}\n"
                f"Пол: {dog_info[3]}\n"
                f"Порода: {dog_info[4]}\n"
                f"ID владельца: {dog_info[5]}"
            )
            self.dog_info.setText(info_text)

            # Загружаем изображение собаки на основе id
            dog_id = dog_info[0]  # Извлекаем id собаки
            photo_path = f"photo_dog{dog_id}.jpg"  # Путь к файлу фотографии
            if os.path.exists(photo_path):
                pixmap = QPixmap(photo_path)

                # Изменяем размер изображения до 150x150 пикселей
                scaled_pixmap = pixmap.scaled(300, 300, Qt.AspectRatioMode.KeepAspectRatio,
                                              Qt.TransformationMode.SmoothTransformation)
                self.dog_image_label.setPixmap(scaled_pixmap)
            else:
                self.dog_image_label.setText("Фото недоступно")


class CreateShowWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Создать выставку")
        self.setGeometry(350, 350, 400, 400)
        self.layout = QVBoxLayout()

        # Поля для ввода данных
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Название выставки")
        self.layout.addWidget(self.name_input)

        self.date_input = QDateEdit()
        self.date_input.setCalendarPopup(True)
        self.layout.addWidget(self.date_input)

        self.location_input = QLineEdit()
        self.location_input.setPlaceholderText("Местоположение")
        self.layout.addWidget(self.location_input)

        self.organizer_input = QLineEdit()
        self.organizer_input.setPlaceholderText("ID Организатора")
        self.layout.addWidget(self.organizer_input)

        self.judge_input = QLineEdit()
        self.judge_input.setPlaceholderText("ID Судей")
        self.layout.addWidget(self.judge_input)

        # Список для выбора собак, которые будут участвовать
        self.dog_list = QListWidget()
        self.dog_list.setSelectionMode(QListWidget.SelectionMode.MultiSelection)
        self.layout.addWidget(self.dog_list)

        # Загрузка списка собак
        cursor.execute("SELECT id, name FROM Dog")
        for row in cursor.fetchall():
            item = QListWidgetItem(f"{row[0]} - {row[1]}")
            self.dog_list.addItem(item)

        # Кнопка для сохранения данных
        btn_save = QPushButton("Сохранить")
        btn_save.clicked.connect(self.save_show)
        self.layout.addWidget(btn_save)

        self.setLayout(self.layout)

    def export_show_to_excel(self, show_id):
        # Создаем новую книгу и лист
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Выставка"

        # Устанавливаем заголовки без колонки "Оценка"
        headers = ["ID", "Имя собаки", "Порода", "Возраст", "Пол"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Получаем данные из базы данных, исключив колонку оценки и связь с Result
        cursor.execute("""
            SELECT Dog.id, Dog.name, Breed.name, Dog.age, Dog.gender
            FROM Dog 
            JOIN Breed ON Dog.breed_id = Breed.id
            WHERE Dog.id IN (
                SELECT dog_id FROM Result WHERE Result.show_id = ?
            )
        """, (show_id,))
        dogs = cursor.fetchall()

        # Заполняем лист данными
        for row_num, dog_data in enumerate(dogs, 2):
            for col_num, value in enumerate(dog_data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)

        # Указываем путь для сохранения файла
        excel_path = "выставка.xlsx"
        workbook.save(excel_path)

        # Открываем Excel файл
        os.startfile(excel_path)

    def save_show(self):
        name = self.name_input.text()
        date = self.date_input.text()
        location = self.location_input.text()
        organizer_id = self.organizer_input.text()
        judge_id = self.judge_input.text()

        try:
            # Вставка данных о новом шоу
            cursor.execute("INSERT INTO Show (name, date, location, organizer_id) VALUES (?, ?, ?, ?)",
                           (name, date, location, organizer_id))
            conn.commit()

            # Получаем id созданного шоу
            show_id = cursor.lastrowid

            # Добавление выбранных собак в шоу
            selected_dogs = self.dog_list.selectedItems()
            for dog_item in selected_dogs:
                dog_id = int(dog_item.text().split(' - ')[0])  # Разделяем строку, чтобы получить ID собаки
                cursor.execute("INSERT INTO Result (show_id, dog_id, score) VALUES (?, ?, ?)",
                               (show_id, dog_id, 0))

            conn.commit()
            QMessageBox.information(self, "Успех", "Выставка успешно создана!")

            # Экспорт шоу в Excel
            self.export_show_to_excel(show_id)

            self.close()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Ошибка", "Ошибка при создании выставки!")


class ShowListWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Список выставок")
        self.setGeometry(300, 300, 400, 300)

        layout = QVBoxLayout()

        # Создаем элемент для отображения списка выставок
        self.show_list = QListWidget()
        self.load_shows()
        layout.addWidget(self.show_list)

        self.setLayout(layout)

    def load_shows(self):
        """Загружает список выставок из базы данных."""
        self.show_list.clear()
        cursor.execute("SELECT name, date, location FROM Show")

        for row in cursor.fetchall():
            self.show_list.addItem(f"{row[0]} - {row[1]} - {row[2]}")

class EnterResultsWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Выставить результаты")
        self.setGeometry(350, 350, 400, 400)
        self.layout = QVBoxLayout()

        # Поля для выбора выставки и собак
        self.show_combo = QComboBox()
        self.layout.addWidget(QLabel("Выберите выставку:"))
        self.populate_shows()
        self.show_combo.currentIndexChanged.connect(self.populate_dogs)  # Обработчик изменения выбора выставки
        self.layout.addWidget(self.show_combo)

        self.dog_combo = QComboBox()
        self.layout.addWidget(QLabel("Выберите собаку:"))
        self.layout.addWidget(self.dog_combo)

        # Поле для ввода оценки
        self.score_input = QLineEdit()
        self.score_input.setPlaceholderText("Введите оценку")
        self.layout.addWidget(self.score_input)

        # Кнопка для сохранения результата
        btn_save = QPushButton("Сохранить результат")
        btn_save.clicked.connect(self.save_result)
        self.layout.addWidget(btn_save)

        self.setLayout(self.layout)

        # Изначально заполняем список собак для выбранного шоу
        self.populate_dogs()

    def populate_shows(self):
        """Загружаем список выставок."""
        cursor.execute("SELECT id, name FROM Show")
        for row in cursor.fetchall():
            self.show_combo.addItem(row[1], row[0])

    def populate_dogs(self):
        """Загружаем список собак, участвующих в выбранной выставке."""
        show_id = self.show_combo.currentData()
        self.dog_combo.clear()

        cursor.execute("""
            SELECT Dog.id, Dog.name FROM Dog
            JOIN Result ON Dog.id = Result.dog_id
            WHERE Result.show_id = ?
        """, (show_id,))

        dogs = cursor.fetchall()

        for dog in dogs:
            self.dog_combo.addItem(dog[1], dog[0])

    def save_result(self):
        """Сохраняем введенные результаты в базу данных."""
        show_id = self.show_combo.currentData()
        dog_id = self.dog_combo.currentData()
        score = self.score_input.text()

        try:
            # Вставляем новую запись с результатом, вместо обновления
            cursor.execute("""
                INSERT INTO Result (show_id, dog_id, score) VALUES (?, ?, ?)
            """, (show_id, dog_id, score))
            conn.commit()
            QMessageBox.information(self, "Успех", "Результат успешно сохранен!")
            self.close()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Ошибка", "Произошла ошибка при сохранении результата.")


class ShowResultsWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Результаты выставок")
        self.setGeometry(300, 300, 400, 300)

        layout = QVBoxLayout()

        # Добавляем выпадающее меню для выбора выставки
        self.show_filter = QComboBox()
        self.show_filter.addItem("Все выставки")
        layout.addWidget(QLabel("Фильтр по выставке:"))
        self.populate_shows()
        self.show_filter.currentIndexChanged.connect(self.load_results)  # Обработчик изменения выбора выставки
        layout.addWidget(self.show_filter)

        # Создаем элемент для отображения результатов выставок
        self.results_list = QListWidget()
        layout.addWidget(self.results_list)

        # Область "Итого"
        self.summary_label = QLabel("Итого:")
        layout.addWidget(self.summary_label)

        self.winner_label = QLabel("Победитель:")
        layout.addWidget(self.winner_label)

        self.setLayout(layout)

        # Изначально загружаем все результаты
        self.load_results()

    def populate_shows(self):
        """Загружаем список выставок для фильтрации."""
        cursor.execute("SELECT name FROM Show")
        for row in cursor.fetchall():
            self.show_filter.addItem(row[0])

    def load_results(self):
        """Загружаем результаты выставок из базы данных."""
        self.results_list.clear()
        selected_show = self.show_filter.currentText()

        if selected_show == "Все выставки":
            cursor.execute("""
                SELECT Show.name, Dog.name, Result.score FROM Result
                JOIN Show ON Result.show_id = Show.id
                JOIN Dog ON Result.dog_id = Dog.id
            """)
        else:
            cursor.execute("""
                SELECT Dog.name, AVG(Result.score) as avg_score FROM Result
                JOIN Show ON Result.show_id = Show.id
                JOIN Dog ON Result.dog_id = Dog.id
                WHERE Show.name = ?
                GROUP BY Dog.id
            """, (selected_show,))

            results = cursor.fetchall()
            scores = {}

            for row in results:
                dog_name, avg_score = row
                self.results_list.addItem(f"Собака: {dog_name}, Средняя оценка: {avg_score:.2f}")
                scores[dog_name] = avg_score

            if scores:
                # Вычисляем победителя
                winner = max(scores, key=scores.get)
                self.winner_label.setText(f"Победитель: {winner} с оценкой {scores[winner]:.2f}")
            else:
                self.winner_label.setText("Победитель: не определен")

        if selected_show != "Все выставки":
            # Обновить метку "Итого"
            self.summary_label.setText(f"Итого: {len(results)} собак")
        else:
            self.summary_label.setText("Итого:")


# Запуск приложения
app = QApplication(sys.argv)
APP_STYLE = """
    QWidget {
        background-color: #ffffff;
        font-family: Courier New;
        font-size: 16px;
    }

    QLabel {
        color: #333;
    }

    QLineEdit {
        border: 1px solid #ccc;
        border-radius: 5px;
        padding: 8px;
        font-size: 16px;
    }

    QPushButton {
        background-color: #725f9b;
        color: white;
        padding: 10px;
        border: none;
        border-radius: 5px;
        font-size: 16px;
    }

    QPushButton:hover {
        background-color: #544672;
    }

    QPushButton:pressed {
        background-color: #451e99;
    }

    QListWidget {
        border: 1px solid #ccc;
        border-radius: 5px;
        padding: 8px;
    }

    QComboBox {
        border: 1px solid #ccc;
        border-radius: 5px;
        padding: 5px;
    }
"""

# Применяем стиль к приложению
setup_database()
app.setStyleSheet(APP_STYLE)
window = AuthWindow()
window.show()
sys.exit(app.exec())